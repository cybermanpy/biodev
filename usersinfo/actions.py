# coding=utf-8

from django.http import HttpResponse
import xlwt
import csv
import openpyxl
from django.utils.encoding import smart_str #csv
from openpyxl.utils import get_column_letter #xlsx
from openpyxl.styles import Font

def export_xls(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=users.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet("users")
    
    row_num = 0
    
    columns = [
        (u"ID", 2000),
        (u"Nombre", 6000),
        (u"Cedula", 8000),
    ]

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    for col_num in xrange(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        # set column width
        ws.col(col_num).width = columns[col_num][1]

    font_style = xlwt.XFStyle()
    font_style.alignment.wrap = 1
    
    for obj in queryset:
        row_num += 1
        row = [
            obj.userid,
            obj.name,
            obj.ssn,
        ]
        for col_num in xrange(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
            
    wb.save(response)
    return response
    
export_xls.short_description = u"Export XLS"


def export_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=users.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"ID"),
        smart_str(u"Nombre"),
        smart_str(u"Cedula"),
    ])
    for obj in queryset:
        writer.writerow([
            smart_str(obj.userid),
            smart_str(obj.name),
            smart_str(obj.ssn),
        ])
    return response
export_csv.short_description = u"Export CSV"


def export_xlsx(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=users.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "users"

    row_num = 0

    columns = [
        (u"ID", 40),
        (u"Nombre", 60),
        (u"Cedula", 27),
    ]

    italic24Font = Font(name='Time New Roman', size=18, bold=True)

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.font = italic24Font
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.userid,
            obj.name,
            obj.ssn,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            # c.style.alignment.wrap_text = True

    wb.save(response)
    return response

export_xlsx.short_description = u"Export XLSX"