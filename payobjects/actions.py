# coding=utf-8

from django.http import HttpResponse
import xlwt
import csv
import openpyxl
from django.utils.encoding import smart_str #csv
from openpyxl.utils import get_column_letter #xlsx
from openpyxl.styles import Font

def export_xls(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=pagos.xls'
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet("pagos")
    
    row_num = 0
    
    columns = [
        (u"Linea", 2000),
        (u"Cedula", 2000),
        (u"Nombre", 8000),
        (u"Concepto", 2500),
        (u"Denominación", 8000),
        (u"Monto", 3000),
        (u"Mes", 3000),
        (u"Año", 2000),
    ]

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    for col_num in xrange(len(columns)):
        ws.write(row_num, col_num, columns[col_num][0], font_style)
        # set column width
        ws.col(col_num).width = columns[col_num][1]

    font_style = xlwt.XFStyle()
    font_style.alignment.wrap = 1
    
    for obj in queryset:
        row_num += 1
        row = [
            obj.fkline.number,
            obj.fkuserinfo.ssn,
            obj.fkuserinfo.name,
            obj.fkpayobject.number,
            obj.fkpayobject.concept,
            obj.amount,
            obj.fkmonth.description,
            obj.fkyear.description,
        ]
        for col_num in xrange(len(row)):
            ws.write(row_num, col_num, row[col_num], font_style)
            
    wb.save(response)
    return response
    
export_xls.short_description = u"Export XLS"


def export_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=pagos.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8')) # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"Linea"),
        smart_str(u"Cedula"),
        smart_str(u"Nombre"),
        smart_str(u"Concepto"),
        smart_str(u"Denominación"),
        smart_str(u"Monto"),
        smart_str(u"Mes"),
        smart_str(u"Año"),
    ])
    for obj in queryset:
        writer.writerow([
            smart_str(obj.fkline.number),
            smart_str(obj.fkuserinfo.ssn),
            smart_str(obj.fkuserinfo.name),
            smart_str(obj.fkpayobject.number),
            smart_str(obj.fkpayobject.concept),
            smart_str(obj.amount),
            smart_str(obj.fkmonth.description),
            smart_str(obj.fkyear.description),
        ])
    return response
export_csv.short_description = u"Export CSV"


def export_xlsx(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=pagos.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "pagos"

    row_num = 0

    columns = [
        (u"Linea", 40),
        (u"Cedula", 60),
        (u"Mombre", 27),
        (u"Concepto", 27),
        (u"Denominación", 27),
        (u"Monto", 27),
        (u"Mes", 27),
        (u"Año", 27),
    ]

    italic24Font = Font(name='Time New Roman', size=18, bold=True)

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.font = italic24Font
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.fkline.number,
            obj.fkuserinfo.ssn,
            obj.fkuserinfo.name,
            obj.fkpayobject.number,
            obj.fkpayobject.concept,
            obj.amount,
            obj.fkmonth.description,
            obj.fkyear.description,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            # c.style.alignment.wrap_text = True

    wb.save(response)
    return response

export_xlsx.short_description = u"Export XLSX"